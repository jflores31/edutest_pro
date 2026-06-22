"""Parseo de archivos de importación de alumnos (CSV / XLSX).

Formato de columnas esperado: **DNI, Nombres, Apellidos** (en ese orden, o
identificadas por encabezado). El DNI se guarda como ``Student.code``.

Acepta:
- ``.csv`` / ``.txt`` (delimitador ``,`` o ``;``, con o sin BOM, UTF-8 o Latin-1)
- ``.xlsx`` (vía openpyxl)
"""

import csv
import io

MAX_ROWS = 5000

# Encabezados aceptados (normalizados) -> campo interno
_HEADER_MAP = {
    "dni": "dni", "code": "dni", "codigo": "dni", "código": "dni",
    "documento": "dni", "doc": "dni",
    "nombres": "first_name", "nombre": "first_name",
    "first_name": "first_name", "first name": "first_name",
    "apellidos": "last_name", "apellido": "last_name",
    "last_name": "last_name", "last name": "last_name",
}


class StudentFileError(Exception):
    """Error de formato/lectura del archivo de alumnos."""


def parse_student_file(django_file):
    """Lee un archivo subido y devuelve [{dni, first_name, last_name}, ...]."""
    name = (getattr(django_file, "name", "") or "").lower()
    if name.endswith(".xlsx"):
        rows = _read_xlsx(django_file)
    elif name.endswith(".csv") or name.endswith(".txt"):
        rows = _read_csv(django_file)
    else:
        raise StudentFileError("Formato no soportado. Usa un archivo .csv o .xlsx.")
    return _normalize_rows(rows)


def _read_csv(django_file):
    raw = django_file.read()
    if isinstance(raw, bytes):
        try:
            text = raw.decode("utf-8-sig")  # quita BOM si existe
        except UnicodeDecodeError:
            text = raw.decode("latin-1")
    else:
        text = raw
    sample = text[:2048]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [list(r) for r in reader]


def _read_xlsx(django_file):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise StudentFileError("Soporte para XLSX no disponible en el servidor.")
    try:
        wb = load_workbook(django_file, read_only=True, data_only=True)
    except Exception:
        raise StudentFileError("No se pudo leer el archivo XLSX. ¿Está dañado?")
    try:
        ws = wb.active
        return [
            ["" if c is None else str(c) for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
    finally:
        wb.close()


def _norm(value):
    return str(value or "").strip().lower()


def _cell(row, i):
    try:
        return str(row[i]).strip()
    except (IndexError, TypeError):
        return ""


def _normalize_rows(rows):
    # Descarta filas totalmente vacías
    rows = [r for r in rows if any(_norm(c) for c in r)]
    if not rows:
        return []

    first = [_norm(c) for c in rows[0]]
    has_header = any(c in _HEADER_MAP for c in first)

    if has_header:
        idx = {}
        for i, c in enumerate(first):
            field = _HEADER_MAP.get(c)
            if field and field not in idx:
                idx[field] = i
        missing = [f for f in ("dni", "first_name", "last_name") if f not in idx]
        if missing:
            raise StudentFileError(
                "Encabezados requeridos no encontrados: se esperan columnas "
                "DNI, Nombres y Apellidos."
            )
        data_rows = rows[1:]
        out = [
            {
                "dni": _cell(r, idx["dni"]),
                "first_name": _cell(r, idx["first_name"]),
                "last_name": _cell(r, idx["last_name"]),
            }
            for r in data_rows
        ]
    else:
        # Sin encabezado: se asume el orden DNI, Nombres, Apellidos
        out = [
            {"dni": _cell(r, 0), "first_name": _cell(r, 1), "last_name": _cell(r, 2)}
            for r in rows
        ]

    out = [r for r in out if r["dni"] or r["first_name"] or r["last_name"]]
    if len(out) > MAX_ROWS:
        raise StudentFileError(f"Máximo {MAX_ROWS} alumnos por importación.")
    return out
